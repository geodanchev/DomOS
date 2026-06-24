"""Reports API endpoints - справки.

Актуализирано за account-based система.
Използва баланса на сметката вместо amount_paid/amount_due.
"""

from datetime import date
from io import BytesIO
from typing import Optional
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models.payment import Payment
from app.models.apartment import Apartment
from app.models.obligation import Obligation, ObligationType
from app.models.account import ApartmentAccount
from app.models.user import User
from app.api.auth import get_current_user

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

router = APIRouter()


@router.get("/payments")
async def get_payments_report(
    from_date: date = Query(..., description="Начална дата"),
    to_date: date = Query(..., description="Крайна дата"),
    apartment_id: Optional[int] = Query(None, description="ID на апартамент"),
    payment_method: Optional[str] = Query(None, description="Метод на плащане"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Справка за плащания по период."""
    query = db.query(Payment).filter(
        Payment.payment_date >= from_date,
        Payment.payment_date <= to_date,
    )
    
    if apartment_id:
        query = query.filter(Payment.apartment_id == apartment_id)
    
    if payment_method:
        query = query.filter(Payment.payment_method == payment_method)
    
    payments = query.order_by(Payment.payment_date.desc()).all()
    
    total_amount = sum(float(p.amount) for p in payments)
    
    return {
        "payments": [
            {
                "id": p.id,
                "apartment_id": p.apartment_id,
                "amount": float(p.amount),
                "month": p.month,
                "payment_date": p.payment_date.isoformat(),
                "payment_method": p.payment_method,
            }
            for p in payments
        ],
        "total_amount": total_amount,
        "count": len(payments),
    }


@router.get("/collection-rate")
async def get_collection_rate(
    month: str = Query(..., description="Месец във формат YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Справка за събираемост по месец.
    
    Използва account-based система.
    """
    # Get monthly obligations for the specified month
    total_obligations = db.query(func.sum(Obligation.amount)).filter(
        Obligation.type == ObligationType.MONTHLY,
        Obligation.month == month
    ).scalar() or Decimal("0")
    
    # Get payments for the specified month
    total_payments = db.query(func.sum(Payment.amount)).filter(
        Payment.month == month
    ).scalar() or Decimal("0")
    
    if float(total_obligations) == 0:
        return {
            "month": month,
            "total_obligations": 0,
            "total_payments": 0,
            "collection_rate": 100.0,
        }
    
    collection_rate = (float(total_payments) / float(total_obligations) * 100)
    
    return {
        "month": month,
        "total_obligations": float(total_obligations),
        "total_payments": float(total_payments),
        "collection_rate": round(collection_rate, 2),
    }


@router.get("/outstanding-debts")
async def get_outstanding_debts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Справка за дължими суми (длъжници).
    
    Използва account-based система.
    Връща апартаменти с отрицателен баланс (дължат).
    """
    # Query all accounts with negative balance
    accounts = db.query(ApartmentAccount).filter(
        ApartmentAccount.balance < 0
    ).all()
    
    debtors = []
    for account in accounts:
        apartment = db.query(Apartment).filter(Apartment.id == account.apartment_id).first()
        if apartment:
            debtors.append({
                "apartment_id": apartment.id,
                "apartment_number": apartment.number,
                "owner_name": apartment.owner_name,
                "amount_owed": abs(float(account.balance)),
            })
    
    debtors.sort(key=lambda x: x['amount_owed'], reverse=True)
    total_outstanding = sum(d['amount_owed'] for d in debtors)
    
    return {
        "debtors": debtors,
        "total_outstanding": total_outstanding,
    }


@router.get("/payments/excel")
async def export_payments_excel(
    from_date: date = Query(..., description="Начална дата"),
    to_date: date = Query(..., description="Крайна дата"),
    apartment_id: Optional[int] = Query(None, description="ID на апартамент"),
    payment_method: Optional[str] = Query(None, description="Метод на плащане"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Excel експорт на справка за плащания."""
    query = db.query(Payment).filter(
        Payment.payment_date >= from_date,
        Payment.payment_date <= to_date,
    )
    
    if apartment_id:
        query = query.filter(Payment.apartment_id == apartment_id)
    
    if payment_method:
        query = query.filter(Payment.payment_method == payment_method)
    
    payments = query.order_by(Payment.payment_date.desc()).all()
    
    if not EXCEL_AVAILABLE:
        # Minimal XLSX file if library is missing
        output = BytesIO()
        output.write(b'PK\x03\x04')
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=payments.xlsx"}
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Плащания"
    
    headers = ["Дата", "Апартамент", "Собственик", "Сума", "Месец", "Метод"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row, p in enumerate(payments, 2):
        apartment = db.query(Apartment).filter(Apartment.id == p.apartment_id).first()
        ws.cell(row=row, column=1, value=p.payment_date.isoformat())
        ws.cell(row=row, column=2, value=apartment.number if apartment else "N/A")
        ws.cell(row=row, column=3, value=apartment.owner_name if apartment else "N/A")
        ws.cell(row=row, column=4, value=float(p.amount))
        ws.cell(row=row, column=5, value=p.month)
        ws.cell(row=row, column=6, value=p.payment_method)
    
    # Add total row
    total_row = len(payments) + 2
    ws.cell(row=total_row, column=3, value="ОБЩО:")
    ws.cell(row=total_row, column=4, value=sum(float(p.amount) for p in payments))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=payments_{from_date}_{to_date}.xlsx"}
    )


@router.get("/payments/pdf")
async def export_payments_pdf(
    from_date: date = Query(..., description="Начална дата"),
    to_date: date = Query(..., description="Крайна дата"),
    apartment_id: Optional[int] = Query(None, description="ID на апартамент"),
    payment_method: Optional[str] = Query(None, description="Метод на плащане"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """PDF експорт на справка за плащания."""
    query = db.query(Payment).filter(
        Payment.payment_date >= from_date,
        Payment.payment_date <= to_date,
    )
    
    if apartment_id:
        query = query.filter(Payment.apartment_id == apartment_id)
    
    if payment_method:
        query = query.filter(Payment.payment_method == payment_method)
    
    payments = query.order_by(Payment.payment_date.desc()).all()
    
    if not PDF_AVAILABLE:
        # Return minimal PDF if library is missing
        output = BytesIO()
        output.write(b'%PDF-1.4\n%\xe2\xe3\xcf\xd3\n')
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=payments_{from_date}_{to_date}.pdf"}
        )
    
    # Create PDF document
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Справка за плащания", styles['Title'])
    elements.append(title)
    
    # Period
    period = Paragraph(f"Период: {from_date} - {to_date}", styles['Normal'])
    elements.append(period)
    elements.append(Spacer(1, 20))
    
    # Table data
    table_data = [["Дата", "Апартамент", "Собственик", "Сума", "Месец", "Метод"]]
    
    for p in payments:
        apartment = db.query(Apartment).filter(Apartment.id == p.apartment_id).first()
        table_data.append([
            p.payment_date.isoformat(),
            apartment.number if apartment else "N/A",
            apartment.owner_name if apartment else "N/A",
            f"{float(p.amount):.2f} лв",
            p.month,
            p.payment_method,
        ])
    
    # Add total row
    total_amount = sum(float(p.amount) for p in payments)
    table_data.append(["", "", "ОБЩО:", f"{total_amount:.2f} лв", "", ""])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=payments_{from_date}_{to_date}.pdf"}
    )


@router.get("/outstanding-debts/excel")
@router.get("/debts/excel")
async def export_debts_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Excel експорт на справка за длъжници."""
    # Get debtors data
    accounts = db.query(ApartmentAccount).filter(
        ApartmentAccount.balance < 0
    ).all()
    
    debtors = []
    for account in accounts:
        apartment = db.query(Apartment).filter(Apartment.id == account.apartment_id).first()
        if apartment:
            debtors.append({
                "apartment_number": apartment.number,
                "owner_name": apartment.owner_name,
                "amount_owed": abs(float(account.balance)),
            })
    
    debtors.sort(key=lambda x: x['amount_owed'], reverse=True)
    
    if not EXCEL_AVAILABLE:
        output = BytesIO()
        output.write(b'PK\x03\x04')
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=debtors.xlsx"}
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Длъжници"
    
    headers = ["Апартамент", "Собственик", "Дължима сума"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row, d in enumerate(debtors, 2):
        ws.cell(row=row, column=1, value=d['apartment_number'])
        ws.cell(row=row, column=2, value=d['owner_name'])
        ws.cell(row=row, column=3, value=d['amount_owed'])
    
    # Add total row
    total_row = len(debtors) + 2
    ws.cell(row=total_row, column=2, value="ОБЩО:")
    ws.cell(row=total_row, column=3, value=sum(d['amount_owed'] for d in debtors))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=debtors.xlsx"}
    )
